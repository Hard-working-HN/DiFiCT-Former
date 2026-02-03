from pathlib import Path
import re
import shutil
from datetime import datetime
import pandas as pd
import numpy as np

SRC_DIR = Path(r"F:\Article2\能源年鉴\xlsx_Files\2022")
SUM_FILE = Path(r"能源统计年鉴汇总-省份.xlsx")
OUT_DIR = Path(r"F:\Article2\能源年鉴\匹配日志-省份")
OUT_DIR.mkdir(parents=True, exist_ok=True)

FILL_ONLY_WHEN_NA = True
WRITE_EMPTY = False

HEADER_LOCS = [
    ("Sheet1_R3", 0, 2),
    ("Sheet1_R43", 0, 42),
    ("Sheet2_R3", 1, 2),
]

def to_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

def rm_spaces(s: str) -> str:
    s = to_str(s)
    return s.replace(" ", "").replace("\u3000", "").replace("\xa0", "").replace("\t", "")

CHINESE_NUMS = "一二三四五六七八九十百千〇零"

def strip_leading_indexing(name: str) -> str:
    s = to_str(name)
    patterns = [
        r"^\s*[\(（]?\d+[\)）\.、\s-]*",
        rf"^\s*[{CHINESE_NUMS}]+[\)）\.、\s-]*",
        r"^\s*[IVXLC]+[\)）\.、\s-]*",
        r"^\s*[①②③④⑤⑥⑦⑧⑨⑩]\s*",
    ]
    for p in patterns:
        s = re.sub(p, "", s)
    return s.strip()

def normalize_industry(s: str) -> str:
    return rm_spaces(strip_leading_indexing(s))

def energy_cn_key(s: str) -> str:
    s = rm_spaces(s)
    keep = []
    for ch in s:
        code = ord(ch)
        if (0x4E00 <= code <= 0x9FFF) or (0x3400 <= code <= 0x4DBF) or (0xF900 <= code <= 0xFAFF):
            keep.append(ch)
    return "".join(keep)

def clean_numberlike(x):
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)
    s = to_str(x)
    if s == "":
        return np.nan
    s = (
        s.replace(",", "")
         .replace("，", "")
         .replace(" ", "")
         .replace("\u3000", "")
         .replace("\t", "")
         .replace("…", "")
    )
    s = re.sub(r"[\*\#\^]+$", "", s)
    s = s.replace("－", "-").replace("–", "-").replace("—", "-")
    try:
        return float(s)
    except:
        return np.nan

def parse_province_year_from_filename(fname: str):
    base = Path(fname).stem
    m = re.search(r"^\s*\d+-\d+\s+(.+?)能源平衡表.*?-(\d{4})\s*$", base)
    if m:
        return m.group(1), m.group(2)
    m2 = re.search(r"(\d{4})", base)
    prov = base
    year = m2.group(1) if m2 else ""
    return prov, year

def load_summary(sum_file: Path):
    xls = pd.ExcelFile(sum_file)
    first_sheet_name = xls.sheet_names[0]
    df = pd.read_excel(sum_file, sheet_name=first_sheet_name, dtype=str)
    if df.shape[1] < 4:
        raise ValueError("Summary file must have at least 4 columns: Province, Year, Industry, and >=1 Energy column.")

    cols = list(df.columns)
    province_col, year_col, industry_col = cols[0], cols[1], cols[2]
    energy_cols = cols[3:]

    df[province_col] = df[province_col].map(to_str)
    df[year_col] = df[year_col].map(to_str)
    df[industry_col] = df[industry_col].map(to_str)
    df["_行业规范"] = df[industry_col].map(normalize_industry)

    cn_to_cols = {}
    for c in energy_cols:
        key = energy_cn_key(c)
        if key:
            cn_to_cols.setdefault(key, set()).add(c)
    energy_unique_map = {k: list(v)[0] for k, v in cn_to_cols.items() if len(v) == 1}
    ambiguous_keys = {k: list(v) for k, v in cn_to_cols.items() if len(v) > 1}

    return (
        df,
        first_sheet_name,
        province_col,
        year_col,
        industry_col,
        energy_cols,
        energy_unique_map,
        ambiguous_keys,
    )

def compute_blocks(sheets, header_locs):
    per_sheet_headers = {}
    for tag, si, hdr in header_locs:
        per_sheet_headers.setdefault(si, []).append((hdr, tag))
    for si in per_sheet_headers:
        per_sheet_headers[si].sort(key=lambda x: x[0])

    blocks = []
    for si, hdrs in per_sheet_headers.items():
        nrow = sheets[si].shape[0] if si < len(sheets) and not sheets[si].empty else 0
        for idx, (hdr, tag) in enumerate(hdrs):
            if nrow == 0 or hdr >= nrow:
                blocks.append((tag, si, hdr, None, None))
                continue
            next_hdr = hdrs[idx + 1][0] if idx + 1 < len(hdrs) else None
            data_end = (next_hdr - 1) if (next_hdr is not None) else (nrow - 1)
            blocks.append((tag, si, hdr, hdr + 1, data_end))
    return blocks

def process_one_workbook(
    xlsx_path: Path,
    sub_sum: pd.DataFrame,
    industry_col: str,
    energy_unique_map: dict,
    ambiguous_keys: dict,
):
    xls = pd.ExcelFile(xlsx_path)
    sheet_names = xls.sheet_names

    sheets = []
    for si in range(2):
        if si < len(sheet_names):
            sheets.append(pd.read_excel(xlsx_path, sheet_name=sheet_names[si], header=None))
        else:
            sheets.append(pd.DataFrame())

    target_inds = set(sub_sum["_行业规范"].map(to_str))
    ambiguous_key_set = set(ambiguous_keys.keys())
    valid_energy_keys = set(energy_unique_map.keys())

    blocks = compute_blocks(sheets, HEADER_LOCS)
    tag_order = {t: i for i, (t, _, _) in enumerate(HEADER_LOCS)}
    blocks.sort(key=lambda b: tag_order.get(b[0], 999))

    decisions = {}
    decided_pairs = set()
    assign_rows = []
    energy_report_rows = []
    industry_report_rows = []
    seen_industry = set()

    for tag, si, hdr_iloc, dstart, dend in blocks:
        df = sheets[si] if si < len(sheets) else pd.DataFrame()
        if df.empty or dstart is None or dend is None or hdr_iloc >= df.shape[0]:
            energy_report_rows.append(
                {
                    "省份表位置": tag,
                    "列索引": None,
                    "省份表_能源原文": "",
                    "能源中文键": "",
                    "是否命中汇总能源列": False,
                    "是否冲突键": False,
                    "命中列名": "",
                    "备注": "工作表缺失/表头越界",
                }
            )
            continue

        header_row = df.iloc[hdr_iloc].tolist()
        header_key_by_col = {}
        for j in range(2, len(header_row)):
            raw = to_str(header_row[j])
            if raw == "":
                continue
            key = energy_cn_key(raw)
            is_conflict = key in ambiguous_key_set
            hit = (key in valid_energy_keys) and (not is_conflict)
            hit_col = energy_unique_map.get(key, "") if hit else ""
            energy_report_rows.append(
                {
                    "省份表位置": tag,
                    "列索引": j,
                    "省份表_能源原文": raw,
                    "能源中文键": key,
                    "是否命中汇总能源列": bool(hit),
                    "是否冲突键": bool(is_conflict),
                    "命中列名": hit_col,
                    "备注": "",
                }
            )
            header_key_by_col[j] = (key, hit, is_conflict, hit_col)

        first_ind_row = None
        for r in range(dstart, min(dend + 1, df.shape[0])):
            ind_raw = to_str(df.iat[r, 0]) if df.shape[1] > 0 else ""
            if ind_raw == "" or ind_raw in ("项目", "Item"):
                continue
            ind_norm = normalize_industry(ind_raw)
            if ind_norm in target_inds:
                first_ind_row = r
                break
        if first_ind_row is None:
            continue

        for r in range(first_ind_row, min(dend + 1, df.shape[0])):
            ind_raw = to_str(df.iat[r, 0]) if df.shape[1] > 0 else ""
            if ind_raw == "" or ind_raw in ("项目", "Item"):
                continue
            ind_norm = normalize_industry(ind_raw)

            if ind_norm not in seen_industry:
                seen_industry.add(ind_norm)
                industry_report_rows.append(
                    {
                        "来源块": tag,
                        "省份表_行业原文": ind_raw,
                        "省份表_行业规范名": ind_norm,
                        "是否命中汇总行业": bool(ind_norm in target_inds),
                    }
                )

            if ind_norm not in target_inds:
                continue

            for j in range(2, df.shape[1]):
                if j not in header_key_by_col:
                    continue
                key, hit, is_conflict, hit_col = header_key_by_col[j]
                if (not hit) or is_conflict or (not key):
                    continue

                pair = (ind_norm, hit_col)
                if pair in decided_pairs:
                    continue

                raw_val = df.iat[r, j] if j < df.shape[1] else np.nan
                val = clean_numberlike(raw_val)
                decisions[pair] = val
                decided_pairs.add(pair)

                assign_rows.append(
                    {
                        "选用块": tag,
                        "行业规范名": ind_norm,
                        "能源列名": hit_col,
                        "取值": (None if pd.isna(val) else float(val)),
                        "是否空值": bool(pd.isna(val)),
                        "数据行": r + 1,
                        "表头行": hdr_iloc + 1,
                        "列索引": j,
                    }
                )

    return decisions, pd.DataFrame(energy_report_rows), pd.DataFrame(industry_report_rows), pd.DataFrame(assign_rows)

def main():
    (
        sum_df,
        first_sheet_name,
        province_col,
        year_col,
        industry_col,
        energy_cols,
        energy_unique_map,
        ambiguous_keys,
    ) = load_summary(SUM_FILE)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = SUM_FILE.with_name(f"{SUM_FILE.stem}_backup_{ts}{SUM_FILE.suffix}")
    try:
        shutil.copy2(SUM_FILE, backup_path)
        print(f"[BACKUP] Backup created: {backup_path}")
    except Exception as e:
        print(f"[WARN] Backup failed (continuing): {e}")

    if ambiguous_keys:
        print("[WARN] Found ambiguous Chinese energy keys (same key maps to multiple columns). These keys will be skipped:")
        for k, cols in ambiguous_keys.items():
            print(f"  key={k} -> columns={cols}")

    energy_report_all = []
    industry_report_all = []
    assign_report_all = []

    xlsx_files = sorted(SRC_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"[WARN] No .xlsx files found in directory: {SRC_DIR}")
        return

    for fp in xlsx_files:
        prov, year = parse_province_year_from_filename(fp.name)
        if not prov or not year:
            print(f"[SKIP] Cannot parse province/year from filename: {fp.name}")
            continue

        mask_py = (sum_df[province_col].map(to_str) == to_str(prov)) & (sum_df[year_col].map(to_str) == to_str(year))
        sub_idx = sum_df.index[mask_py]
        if len(sub_idx) == 0:
            print(f"[SKIP] Not found in summary: province={prov}, year={year}")
            continue

        sub_cols = [province_col, year_col, industry_col, "_行业规范"] + energy_cols
        sub_sum = sum_df.loc[sub_idx, sub_cols].copy()

        try:
            decisions, energy_report, industry_report, assign_rows = process_one_workbook(
                fp, sub_sum, industry_col, energy_unique_map, ambiguous_keys
            )
        except Exception as e:
            print(f"[ERROR] Failed processing {fp.name}: {e}")
            continue

        if decisions:
            for row_i in sub_sum.index:
                ind_norm = to_str(sub_sum.at[row_i, "_行业规范"])
                for c in energy_cols:
                    pair = (ind_norm, c)
                    if pair not in decisions:
                        continue
                    val = decisions[pair]
                    if pd.isna(val):
                        if WRITE_EMPTY:
                            sub_sum.at[row_i, c] = np.nan
                        else:
                            continue
                    else:
                        if FILL_ONLY_WHEN_NA:
                            sub_sum.at[row_i, c] = float(val)
                        else:
                            sub_sum.at[row_i, c] = float(val)

        for c in energy_cols:
            write_vals = sub_sum[c]
            if WRITE_EMPTY:
                sum_df.loc[write_vals.index, c] = write_vals
            else:
                mask_write = write_vals.notna()
                sum_df.loc[write_vals.index[mask_write], c] = write_vals[mask_write]

        if not energy_report.empty:
            energy_report.insert(0, "省份", prov)
            energy_report.insert(1, "年份", year)
            energy_report_all.append(energy_report)
        if not industry_report.empty:
            industry_report.insert(0, "省份", prov)
            industry_report.insert(1, "年份", year)
            industry_report_all.append(industry_report)
        if not assign_rows.empty:
            assign_rows.insert(0, "省份", prov)
            assign_rows.insert(1, "年份", year)
            assign_report_all.append(assign_rows)

        print(f"[OK] Processed and written back: {prov} {year} (records={len(assign_rows)})")

    try:
        sum_df[energy_cols] = sum_df[energy_cols].apply(pd.to_numeric, errors="ignore")
    except Exception:
        pass

    try:
        with pd.ExcelWriter(SUM_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sum_df.drop(columns=["_行业规范"]).to_excel(writer, index=False, sheet_name=first_sheet_name)
        print(f"[SAVE] Saved to: {SUM_FILE} (sheet: {first_sheet_name})")
    except TypeError:
        from openpyxl import load_workbook
        wb = load_workbook(SUM_FILE)
        if first_sheet_name in wb.sheetnames:
            ws = wb[first_sheet_name]
            wb.remove(ws)
            wb.save(SUM_FILE)
        with pd.ExcelWriter(SUM_FILE, engine="openpyxl", mode="a") as writer:
            sum_df.drop(columns=["_行业规范"]).to_excel(writer, index=False, sheet_name=first_sheet_name)
        print(f"[SAVE] Saved to: {SUM_FILE} (legacy pandas compatibility, sheet: {first_sheet_name})")

    try:
        rep_path = OUT_DIR / f"匹配报告_首个命中即定_{ts}.xlsx"
        with pd.ExcelWriter(rep_path, engine="xlsxwriter") as writer:
            if energy_report_all:
                pd.concat(energy_report_all, ignore_index=True).to_excel(writer, index=False, sheet_name="表头命中_能源")
            if industry_report_all:
                pd.concat(industry_report_all, ignore_index=True).to_excel(writer, index=False, sheet_name="行业命中")
            if assign_report_all:
                pd.concat(assign_report_all, ignore_index=True).to_excel(writer, index=False, sheet_name="赋值来源_首命中")
        print(f"[LOG] Report saved: {rep_path}")
    except Exception as e:
        print(f"[WARN] Failed to save report: {e}")

if __name__ == "__main__":
    main()
